HOME = 'C:\\Temp\\'
FILENAME = 'test_wb.xlsx'

from openpyxl import Workbook, load_workbook
from random import randint

path = HOME + FILENAME
wb = Workbook()
ws = wb.create_sheet(title = 'random_table')
for i in range(12):
    for j in range(12):
        cell = randint(0, 9)
        ws.cell(column = i+1, row = j+1, value = cell)

wb.save(path)